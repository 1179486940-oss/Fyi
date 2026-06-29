"""
MCP: Excel 操作
生成数据表格报表（Excel），支持样式预设
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from config import get_settings
from utils.logger import get_logger
from utils.helpers import generate_uuid

logger = get_logger(__name__)


class ExcelMCP:
    """Excel 生成 MCP Server"""

    def __init__(self):
        settings = get_settings()
        self._storage_path = settings.storage.storage_path / "reports"
        self._storage_path.mkdir(parents=True, exist_ok=True)
        self._base_url = f"http://{settings.server.host}:{settings.server.port}/files"

        # BA 预设样式模板
        self._style_presets: dict[str, dict] = {
            "default": {"font_size": 11, "header_color": "4472C4", "border": True},
            "compact": {"font_size": 10, "header_color": "2F5496", "border": True},
            "colorful": {"font_size": 12, "header_color": "C55A11", "border": True},
        }

    # ── Excel 生成 ───────────────────────────────────────

    async def generate_excel(
        self,
        data: list[dict],
        sheet_name: str = "Sheet1",
        title: str = "数据报表",
        style_preset: str = "default",
        extra_sheets: Optional[dict[str, list[dict]]] = None,
    ) -> dict[str, Any]:
        """
        生成 Excel 文件
        Args:
            data: 主数据表
            sheet_name: Sheet 名称
            title: 报表标题
            style_preset: 样式预设 (default / compact / colorful)
            extra_sheets: 额外 Sheet {"Sheet名": [数据]}
        Returns:
            {"status": "success", "file_id": "...", "download_url": "..."}
        """
        if not data:
            return {"status": "error", "message": "No data provided"}

        file_id = generate_uuid()
        file_path = self._storage_path / f"{file_id}.xlsx"

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

            wb = openpyxl.Workbook()
            style = self._style_presets.get(style_preset, self._style_presets["default"])

            # ── 主 Sheet ──────────────────────────────
            ws = wb.active
            ws.title = sheet_name
            self._write_sheet(ws, data, title, style)

            # ── 额外 Sheet ────────────────────────────
            if extra_sheets:
                for s_name, s_data in extra_sheets.items():
                    ws2 = wb.create_sheet(title=s_name)
                    self._write_sheet(ws2, s_data, s_name, style)

            wb.save(str(file_path))

            logger.info("Excel generated: %s (%d rows, style=%s)",
                         title, len(data), style_preset)

            return {
                "status": "success",
                "file_id": file_id,
                "download_url": f"{self._base_url}/{file_id}/download",
                "format": "xlsx",
                "row_count": len(data),
            }

        except Exception as e:
            logger.error("Excel generation failed: %s", e)
            return {"status": "error", "message": str(e)}

    def _write_sheet(self, ws, data: list[dict], title: str, style: dict) -> None:
        """向 Sheet 写入数据并应用样式"""
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        if not data:
            ws.cell(row=1, column=1, value="(无数据)")
            return

        # 标题行
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="FFFFFF" if style["header_color"] else "000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data[0]))

        # 表头
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color=style["header_color"], end_color=style["header_color"], fill_type="solid")
        header_font = Font(size=style["font_size"], bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        data_font = Font(size=style["font_size"])
        for row_idx, row in enumerate(data, 4):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
                cell.font = data_font

        # 列宽自适应
        for col_idx, header in enumerate(headers, 1):
            max_len = max(len(str(header)), max((len(str(r.get(header, ""))) for r in data), default=0))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # ── 预设管理 ───────────────────────────────────────

    def add_style_preset(self, name: str, config: dict) -> None:
        """BA 添加新的样式预设"""
        self._style_presets[name] = config
        logger.info("Style preset added: %s", name)

    def list_presets(self) -> list[str]:
        """列出所有可用的样式预设"""
        return list(self._style_presets.keys())
